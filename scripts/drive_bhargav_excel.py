"""Drive Bhargav day-plan .xlsm via Excel COM. Windows + Excel + LAN only."""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSM = (
    ROOT
    / "docs/planning-redesign/planning-comparion/grab-and-go/Copy of Master Template - DAY PLANNING TEMPLATE - 19.08.2026- V1.xlsm"
)
EVIDENCE = ROOT / "docs/planning-compare/runs"


def _excel():
    import win32com.client  # type: ignore

    app = win32com.client.DispatchEx("Excel.Application")
    app.Visible = True
    app.DisplayAlerts = False
    app.AskToUpdateLinks = False
    app.AutomationSecurity = 1  # msoAutomationSecurityLow — COM defaults to macros off
    return app


def _open_copy_readonly(app, xlsm: Path):
    """Never open the master. Copy, then Open ReadOnly. Never Save."""
    scratch = Path(tempfile.gettempdir()) / f"bhargav-scratch-{os.getpid()}.xlsm"
    shutil.copy2(xlsm, scratch)
    wb = app.Workbooks.Open(
        str(scratch),
        ReadOnly=True,
        UpdateLinks=0,
        IgnoreReadOnlyRecommended=True,
    )
    return wb, scratch


def _sheet(wb, name: str):
    want = name.upper()
    for sh in wb.Worksheets:
        if sh.Name.upper() == want:
            return sh
    raise SystemExit(f"missing sheet {name!r}: {[s.Name for s in wb.Worksheets]}")


def _dump_shapes(ws) -> list[dict]:
    out = []
    try:
        for i in range(1, ws.Shapes.Count + 1):
            sh = ws.Shapes.Item(i)
            rec = {"name": sh.Name, "on_action": ""}
            try:
                rec["on_action"] = str(sh.OnAction or "")
            except Exception:
                pass
            out.append(rec)
    except Exception as exc:
        out.append({"error": str(exc)})
    return out


def _click_next(ws) -> str:
    names = ["RefreshStep01", "Button12_Click"]
    last_err = None
    for name in names:
        try:
            ws.Application.Run(name)
            return name
        except Exception as exc:
            last_err = exc
    shapes = _dump_shapes(ws)
    raise SystemExit(f"macro run failed {last_err}; shapes={shapes}")


def _used_rows(ws, col: int, start: int, end: int) -> int:
    last = start
    for r in range(start, end + 1):
        if ws.Cells(r, col).Value not in (None, "", "-"):
            last = r
    return last


def dump_fresh(ws, dest: Path) -> list[dict]:
    from openpyxl import Workbook

    rows = []
    last = _used_rows(ws, 2, 4, 400)
    book = Workbook()
    out = book.active
    out.title = "Fresh Products By Supplier"
    headers = ["row", "B", "C", "D", "E", "F", "G"]
    out.append(headers)
    for r in range(4, last + 1):
        rec = {
            "row": r,
            "B": ws.Cells(r, 2).Value,
            "C": ws.Cells(r, 3).Value,
            "D": ws.Cells(r, 4).Value,
            "E": ws.Cells(r, 5).Value,
            "F": ws.Cells(r, 6).Value,
            "G": ws.Cells(r, 7).Value,
        }
        rows.append(rec)
        out.append([rec["row"], rec["B"], rec["C"], rec["D"], rec["E"], rec["F"], rec["G"]])
    dest.parent.mkdir(parents=True, exist_ok=True)
    book.save(dest)
    return rows


def _write_compare_source(dest: Path, *, code, name, cases, rows):
    from openpyxl import Workbook

    book = Workbook()
    pack = book.active
    pack.title = "PACKING PLAN"
    pack.cell(6, 2, code)
    pack.cell(6, 3, name)
    pack.cell(6, 13, cases)
    fresh = book.create_sheet("Fresh Products By Supplier")
    for rec in rows:
        r = rec["row"]
        fresh.cell(r, 2, rec["B"])
        fresh.cell(r, 3, rec["C"])
        fresh.cell(r, 4, rec["D"])
        fresh.cell(r, 5, rec["E"])
        fresh.cell(r, 6, rec["F"])
        fresh.cell(r, 7, rec["G"])
    dest.parent.mkdir(parents=True, exist_ok=True)
    book.save(dest)


def run(
    *,
    xlsm: Path,
    code: str,
    cases: float,
    plan_date: str,
    wait_sec: int,
    probe_only: bool,
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = EVIDENCE / f"{code}-{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)
    meta = {
        "code": code,
        "cases": cases,
        "plan_date": plan_date,
        "xlsm": str(xlsm),
        "opened_readonly": True,
        "macros": "AutomationSecurity=Low",
        "started": datetime.now().isoformat(timespec="seconds"),
    }

    app = _excel()
    wb = None
    scratch = None
    try:
        wb, scratch = _open_copy_readonly(app, xlsm)
        meta["sheets"] = [s.Name for s in wb.Worksheets]
        meta["opened"] = "copy_readonly"
        home = _sheet(wb, "Home")
        home.Range("C4").Value = plan_date
        meta["home_c4"] = str(home.Range("C4").Value)
        meta["home_c7"] = str(home.Range("C7").Value)

        pack = _sheet(wb, "PACKING PLAN")
        hit_row = None
        for r in range(5, 400):
            cell = pack.Cells(r, 2).Value
            if cell is None:
                continue
            if str(cell).strip().upper() == code.upper():
                hit_row = r
                break
        if hit_row is None:
            raise SystemExit(f"{code} not in PACKING PLAN col B")
        pack.Cells(hit_row, 13).Value = cases
        meta["packing_row"] = hit_row
        meta["packing_name"] = str(pack.Cells(hit_row, 3).Value)
        meta["packing_shapes"] = _dump_shapes(pack)

        (run_dir / "probe.json").write_text(json.dumps(meta, indent=2, default=str))
        if probe_only:
            print(json.dumps(meta, indent=2, default=str))
            return run_dir

        action = _click_next(pack)
        meta["next_step"] = action
        time.sleep(wait_sec)

        fresh = None
        for sh in wb.Worksheets:
            if "FRESH PRODUCTS" in sh.Name.upper() and "SUPPLIER" in sh.Name.upper():
                fresh = sh
                break
        if fresh is None:
            raise SystemExit(f"no Fresh Products By Supplier: {meta['sheets']}")

        dump_xlsx = run_dir / "fresh-products.xlsx"
        rows = dump_fresh(fresh, dump_xlsx)
        (run_dir / "fresh-products.json").write_text(
            json.dumps(rows, indent=2, default=str)
        )
        compare_xlsx = run_dir / "compare-source.xlsx"
        _write_compare_source(
            compare_xlsx,
            code=code,
            name=meta["packing_name"],
            cases=cases,
            rows=rows,
        )
        meta["fresh_rows"] = len(rows)
        meta["dump_xlsx"] = str(dump_xlsx)
        meta["compare_xlsx"] = str(compare_xlsx)
        meta["finished"] = datetime.now().isoformat(timespec="seconds")
        (run_dir / "meta.json").write_text(json.dumps(meta, indent=2, default=str))
        print(run_dir)
        return run_dir
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        app.Quit()
        if scratch is not None:
            scratch.unlink(missing_ok=True)


def run_batch(*, xlsm: Path, codes: list[str], cases: float, plan_date: str, wait_sec: int):
    """One Excel process; each SKU zeros other cases then RefreshStep01."""
    app = _excel()
    wb = None
    scratch = None
    dirs = []
    try:
        wb, scratch = _open_copy_readonly(app, xlsm)
        home = _sheet(wb, "Home")
        home.Range("C4").Value = plan_date
        pack = _sheet(wb, "PACKING PLAN")
        index = {}
        for r in range(5, 400):
            cell = pack.Cells(r, 2).Value
            if cell:
                index[str(cell).strip().upper()] = r
        missing = [c for c in codes if c.upper() not in index]
        if missing:
            raise SystemExit(f"not in PACKING PLAN col B: {missing}")

        prev_row = None
        for code in codes:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            run_dir = EVIDENCE / f"{code}-{stamp}"
            run_dir.mkdir(parents=True, exist_ok=True)
            hit_row = index[code.upper()]
            # Only unlocked Case cells (M). Zeroing every row hits locked cells.
            if prev_row is not None:
                pack.Cells(prev_row, 13).Value = None
            pack.Cells(hit_row, 13).Value = cases
            prev_row = hit_row
            meta = {
                "code": code,
                "cases": cases,
                "plan_date": plan_date,
                "packing_row": hit_row,
                "packing_name": str(pack.Cells(hit_row, 3).Value),
                "started": datetime.now().isoformat(timespec="seconds"),
            }
            action = _click_next(pack)
            meta["next_step"] = action
            time.sleep(wait_sec)
            fresh = None
            for sh in wb.Worksheets:
                if "FRESH PRODUCTS" in sh.Name.upper() and "SUPPLIER" in sh.Name.upper():
                    fresh = sh
                    break
            if fresh is None:
                raise SystemExit("no Fresh Products By Supplier")
            dump_xlsx = run_dir / "fresh-products.xlsx"
            rows = dump_fresh(fresh, dump_xlsx)
            (run_dir / "fresh-products.json").write_text(
                json.dumps(rows, indent=2, default=str)
            )
            compare_xlsx = run_dir / "compare-source.xlsx"
            _write_compare_source(
                compare_xlsx,
                code=code,
                name=meta["packing_name"],
                cases=cases,
                rows=rows,
            )
            meta["fresh_rows"] = len(rows)
            meta["compare_xlsx"] = str(compare_xlsx)
            meta["finished"] = datetime.now().isoformat(timespec="seconds")
            (run_dir / "meta.json").write_text(json.dumps(meta, indent=2, default=str))
            print(run_dir)
            dirs.append(run_dir)
        return dirs
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        app.Quit()
        if scratch is not None:
            scratch.unlink(missing_ok=True)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--code", default="CCKATB1-R6TTA")
    p.add_argument("--codes", default="", help="comma-separated SKUs; one Excel session")
    p.add_argument("--cases", type=float, default=2000)
    p.add_argument("--date", default="24/08/2026")
    p.add_argument("--excel", type=Path, default=DEFAULT_XLSM)
    p.add_argument("--wait", type=int, default=35)
    p.add_argument("--probe-only", action="store_true")
    args = p.parse_args()
    if not args.excel.exists():
        sys.exit(f"missing {args.excel}")
    codes = [c.strip() for c in args.codes.split(",") if c.strip()]
    if codes:
        run_batch(
            xlsm=args.excel,
            codes=codes,
            cases=args.cases,
            plan_date=args.date,
            wait_sec=args.wait,
        )
        return
    run(
        xlsm=args.excel,
        code=args.code,
        cases=args.cases,
        plan_date=args.date,
        wait_sec=args.wait,
        probe_only=args.probe_only,
    )


if __name__ == "__main__":
    main()
