import openpyxl

wb = openpyxl.load_workbook(r"recipe/docs/master - poducrt.xlsx", read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print("---", s)
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(i, row[:15] if row else None)
    n = 0
    for _ in ws.iter_rows(values_only=True):
        n += 1
    print("rows", n)
