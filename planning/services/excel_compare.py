"""Parse a Gazebo day-plan workbook, explode a draft plan, compare RM kg."""

from __future__ import annotations

import json
import re
import warnings
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Sum
from openpyxl import load_workbook

from locations.models import Location
from planning.models import (
    ExcelCompareReport,
    Plan,
    PlanLine,
    PlanLineSource,
    PlanRequirement,
    PlanStatus,
)
from planning.services import explode, lifecycle
from product.models import Product
from recipe.models import Recipe

PACK_SHEET = 'PACKING PLAN'
NAME_MATCH_MIN = 0.84
DEFAULT_MAP = Path('docs/planning-compare/code-map.json')


class ExcelCompareError(Exception):
    pass


def _dec(value) -> Decimal | None:
    if value in (None, '', '-', 0, 0.0, '0'):
        return None
    try:
        qty = Decimal(str(value).strip().replace(',', ''))
    except (InvalidOperation, AttributeError, TypeError):
        return None
    return qty if qty > 0 else None


def _norm_code(value) -> str:
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


def _norm_name(value) -> str:
    text = re.sub(r'\([^)]*\)', ' ', str(value or '').lower())
    return ' '.join(re.sub(r'[^a-z0-9]+', ' ', text).split())


def _cell(row, idx):
    return row[idx] if idx < len(row) else None


def _to_kg(qty: Decimal, unit_name: str | None) -> Decimal | None:
    unit = (unit_name or '').strip().lower()
    if unit in {'g', 'gram', 'grams', 'gm', 'gms'}:
        return qty / Decimal('1000')
    if unit in {'kg', 'kgs', 'kilogram', 'kilograms'}:
        return qty
    if unit in {'mg', 'milligram', 'milligrams'}:
        return qty / Decimal('1000000')
    return None


def _qty(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, 'f')
    return str(value)


def _g(kg: Decimal | None) -> Decimal | None:
    if kg is None:
        return None
    return kg * Decimal('1000')


def load_code_map(path: Path | None = None) -> dict[str, str]:
    target = path or (Path(settings.BASE_DIR) / DEFAULT_MAP)
    if not target.exists():
        return {}
    raw = json.loads(target.read_text())
    return {str(k).strip().upper(): str(v).strip() for k, v in raw.items() if k and v}


def _open_workbook(source):
    warnings.simplefilter('ignore')
    if hasattr(source, 'read'):
        data = source.read()
        if hasattr(source, 'seek'):
            source.seek(0)
        return load_workbook(BytesIO(data), data_only=True, read_only=True)
    return load_workbook(source, data_only=True, read_only=True)


def _sheet_named(wb, exact: str) -> str | None:
    want = exact.upper()
    for name in wb.sheetnames:
        if name.upper() == want:
            return name
    return None


def _sheets_containing(wb, *needles: str) -> list[str]:
    hits = []
    for name in wb.sheetnames:
        upper = name.upper()
        if all(n.upper() in upper for n in needles):
            hits.append(name)
    return hits


def parse_packing(ws) -> list[dict]:
    lines = []
    for i, row in enumerate(ws.iter_rows(min_row=5, max_col=16, values_only=True), 5):
        code = str(_cell(row, 1) or '').strip()
        name = str(_cell(row, 2) or '').strip()
        if not code or code in {'-', 'Product Code'} or name.upper() == 'TOTAL':
            continue
        cases = _dec(_cell(row, 12))
        if cases is None:
            continue
        trays = _dec(_cell(row, 9))
        per_case = _dec(_cell(row, 14))
        packs = trays or (cases * per_case if per_case else cases)
        lines.append({
            'excel_row': i,
            'code': code,
            'name': name,
            'cases': cases,
            'packs': packs,
        })
    return lines


def parse_rm(ws, sheet: str) -> list[dict]:
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=5, max_col=7, values_only=True), 5):
        code = str(_cell(row, 1) or '').strip()
        name = str(_cell(row, 2) or '').strip()
        if not name or name.lower().startswith('grand total'):
            continue
        kg = _dec(_cell(row, 5))
        if kg is None:
            continue
        if code in {'-', 'Purchase Code', '`'}:
            code = ''
        rows.append({
            'excel_row': i,
            'sheet': sheet,
            'code': code,
            'name': name,
            'kg': kg,
        })
    return rows


def parse_workbook(source) -> tuple[list[dict], list[dict]]:
    wb = _open_workbook(source)
    try:
        pack_name = _sheet_named(wb, PACK_SHEET)
        if pack_name is None:
            raise ExcelCompareError(
                f'missing sheet {PACK_SHEET!r}. Found: {wb.sheetnames}'
            )
        fg_lines = parse_packing(wb[pack_name])
        excel_rm = []
        for sheet in _sheets_containing(wb, 'FRESH PRODUCTS'):
            excel_rm.extend(parse_rm(wb[sheet], sheet))
    finally:
        wb.close()
    if not fg_lines:
        raise ExcelCompareError('no PACKING PLAN rows with col M cases > 0')
    return fg_lines, excel_rm


def _product_index(code_map: dict[str, str]):
    products = list(Product.objects.select_related('unit'))
    by_code: dict[str, Product] = {}
    for product in products:
        for raw in (
            product.recipe_code,
            product.alternate_recipe_code,
            product.gff_code,
            str(product.id),
        ):
            key = _norm_code(raw)
            if key:
                by_code.setdefault(key, product)
                if key.endswith('ST') and key.startswith('GFF') and len(key) > 5:
                    by_code.setdefault(key[:-2], product)
    names = [
        (_norm_name(n), p)
        for p in products
        for n in (p.name, p.alternate_name)
        if n
    ]

    def resolve(code: str, name: str):
        mapped = code_map.get((code or '').upper())
        if mapped:
            hit = by_code.get(_norm_code(mapped))
            if hit:
                return hit, 'map'
            try:
                hit = Product.objects.filter(pk=int(mapped)).first()
            except ValueError:
                hit = None
            if hit:
                return hit, 'map-id'
        key = _norm_code(code)
        hit = by_code.get(key)
        if hit:
            return hit, 'code'
        if key.endswith('ST') and key.startswith('GFF'):
            hit = by_code.get(key[:-2])
            if hit:
                return hit, 'gff'
        target = _norm_name(name)
        if not target:
            return None, 'unmapped'
        best, score = None, 0.0
        for nname, product in names:
            ratio = SequenceMatcher(None, target, nname).ratio()
            if ratio > score:
                best, score = product, ratio
        if best is not None and score >= NAME_MATCH_MIN:
            return best, f'name:{score:.2f}'
        return None, 'unmapped'

    return resolve


def _ensure_draft(*, plan_id, plan_date, location, remarks):
    if plan_id:
        plan = Plan.objects.filter(pk=plan_id).first()
        if plan is None:
            raise ExcelCompareError(f'plan {plan_id} not found')
        if plan.status != PlanStatus.DRAFT:
            raise ExcelCompareError('reuse plan must be draft')
        return plan
    existing = Plan.objects.filter(plan_date=plan_date, location=location).first()
    if existing:
        if existing.status != PlanStatus.DRAFT:
            raise ExcelCompareError(
                f'plan {existing.id} already {existing.status} for that date'
            )
        existing.remarks = remarks
        existing.save(update_fields=['remarks', 'updated_at'])
        return existing
    return lifecycle.create_plan(
        plan_date=plan_date,
        location_id=location.id,
        remarks=remarks,
    )


def _system_rm_kg(run_id: int) -> dict[int, dict]:
    cooked = set(Recipe.objects.values_list('product_id', flat=True))
    rows = (
        PlanRequirement.objects.filter(run_id=run_id)
        .values('product_id')
        .annotate(net=Sum('net_required'), gross=Sum('gross_required'))
    )
    products = {
        p.id: p
        for p in Product.objects.filter(
            pk__in=[r['product_id'] for r in rows],
        ).select_related('unit')
    }
    out = {}
    for row in rows:
        product = products[row['product_id']]
        if product.id in cooked:
            continue
        unit = product.unit.name if product.unit_id else ''
        out[product.id] = {
            'product': product,
            'net': row['net'],
            'gross': row['gross'],
            'kg': _to_kg(row['gross'], unit),
            'unit': unit,
        }
    return out


def _fix_note(*, product, sys, excel_kg, dry_run):
    if product is None:
        return 'unmapped — add excel code to code-map.json'
    if sys is None:
        return (
            None
            if dry_run
            else 'not in explode (missing BOM, or treated as intermediate because it has a recipe)'
        )
    if sys['kg'] is None:
        return f'system unit {sys["unit"]!r} is not g/kg'
    diff = sys['kg'] - excel_kg
    pct = (diff / sys['kg']) * Decimal('100') if sys['kg'] else None
    if pct is not None and abs(pct) > 5:
        if 15 <= abs(pct) <= 25:
            return '~20% — check yield_factor / process_loss'
        return 'recipe qty or yield vs Excel; confirm pack size'
    return None


def _fg_payload(row) -> dict:
    product = row['product']
    return {
        'excel_row': row['excel_row'],
        'excel_code': row['code'],
        'excel_name': row['name'],
        'cases': _qty(row['cases']),
        'packs': _qty(row['packs']),
        'explode_qty': _qty(row['qty']),
        'match': row['how'],
        'product_id': product.id if product else None,
        'recipe_code': product.recipe_code if product else None,
        'product_name': product.name if product else None,
    }


def _rm_payload(row, sys, dry_run) -> dict:
    product = row['product']
    excel_kg = row['kg']
    sys_kg = sys['kg'] if sys else None
    excel_g = _g(excel_kg)
    sys_g = _g(sys_kg)
    diff_kg = diff_g = pct = None
    if sys_kg is not None:
        diff_kg = sys_kg - excel_kg
        if excel_g is not None:
            diff_g = sys_g - excel_g if sys_g is not None else None
        if sys_kg:
            pct = (diff_kg / sys_kg) * Decimal('100')
    return {
        'excel_code': row['code'],
        'excel_name': row['name'],
        'sheet': row.get('sheet'),
        'excel_kg': _qty(excel_kg),
        'excel_g': _qty(excel_g),
        'match': row['how'],
        'product_id': product.id if product else None,
        'recipe_code': product.recipe_code if product else None,
        'product_name': product.name if product else None,
        'system_kg': _qty(sys_kg),
        'system_g': _qty(sys_g),
        'diff_kg': _qty(diff_kg),
        'diff_g': _qty(diff_g),
        'diff_pct': float(round(pct, 2)) if pct is not None else None,
        'fix': _fix_note(
            product=product, sys=sys, excel_kg=excel_kg, dry_run=dry_run,
        ),
    }


def run_excel_compare(
    *,
    source,
    location_id: int,
    plan_date,
    qty_mode: str = 'cases',
    dry_run: bool = False,
    plan_id: int | None = None,
    code_map: dict[str, str] | None = None,
    remarks: str | None = None,
) -> dict:
    location = Location.objects.filter(pk=location_id).first()
    if location is None:
        raise ExcelCompareError(f'location {location_id} not found')
    if qty_mode not in ('packs', 'cases'):
        raise ExcelCompareError('qty_mode must be packs or cases')

    fg_lines, excel_rm = parse_workbook(source)
    resolve = _product_index(code_map if code_map is not None else load_code_map())

    mapped_fg = []
    for line in fg_lines:
        product, how = resolve(line['code'], line['name'])
        qty = line['cases'] if qty_mode == 'cases' else line['packs']
        mapped_fg.append({**line, 'product': product, 'how': how, 'qty': qty})
    mapped_rm = []
    for row in excel_rm:
        product, how = resolve(row['code'], row['name'])
        mapped_rm.append({**row, 'product': product, 'how': how})

    unmapped_fg = [r['code'] for r in mapped_fg if r['product'] is None]
    plan = run = None
    sys_by_id: dict[int, dict] = {}
    if not dry_run:
        if unmapped_fg:
            raise ExcelCompareError(
                'unmapped FG: ' + ', '.join(unmapped_fg)
            )
        plan = _ensure_draft(
            plan_id=plan_id,
            plan_date=plan_date,
            location=location,
            remarks=remarks or 'excel compare',
        )
        PlanLine.objects.filter(plan=plan).delete()
        for i, row in enumerate(mapped_fg, start=1):
            product = row['product']
            PlanLine.objects.create(
                plan=plan,
                product=product,
                quantity=row['qty'],
                unit_id=product.unit_id,
                source=PlanLineSource.MANUAL,
                override_consider_stock=False,
                override_full_batches=False,
                override_align_last_batch=False,
                sort_order=i,
            )
        run = explode.run_explode(plan.id)
        sys_by_id = _system_rm_kg(run.id)

    used = set()
    rm_rows = []
    for row in mapped_rm:
        product = row['product']
        sys = sys_by_id.get(product.id) if product else None
        if product:
            used.add(product.id)
        rm_rows.append(_rm_payload(row, sys, dry_run))

    system_only = []
    if not dry_run:
        for pid, sys in sorted(sys_by_id.items()):
            if pid in used:
                continue
            product = sys['product']
            system_only.append({
                'product_id': pid,
                'recipe_code': product.recipe_code,
                'product_name': product.name,
                'system_kg': _qty(sys['kg'] if sys['kg'] is not None else sys['gross']),
                'unit': sys['unit'],
                'fix': 'in explode, not in Excel — extra BOM or code mismatch',
            })

    payload = {
        'plan_id': plan.id if plan else None,
        'run_id': run.id if run else None,
        'run_status': run.status if run else None,
        'location_id': location.id,
        'plan_date': plan_date.isoformat(),
        'qty_mode': qty_mode,
        'dry_run': dry_run,
        'unmapped_fg': unmapped_fg,
        'finished_goods': [_fg_payload(r) for r in mapped_fg],
        'rm_compare': rm_rows,
        'system_only': system_only,
        '_mapped_fg': mapped_fg,
        '_mapped_rm': mapped_rm,
        '_sys_by_id': sys_by_id,
        '_plan': plan,
        '_run': run,
    }
    return payload


def public_result(result: dict) -> dict:
    return {k: v for k, v in result.items() if not k.startswith('_')}


def persist_report(result: dict, file_name: str = '') -> dict:
    pub = public_result(result)
    row = ExcelCompareReport.objects.create(
        location_id=pub['location_id'],
        plan_date=pub['plan_date'],
        dry_run=pub['dry_run'],
        file_name=(file_name or '')[:255],
        plan_id=pub['plan_id'],
        run_id=pub['run_id'],
        payload=pub,
    )
    pub['report_id'] = row.id
    return pub


def report_summary(row: ExcelCompareReport) -> dict:
    return {
        'id': row.id,
        'created_at': row.created_at.isoformat() if row.created_at else None,
        'location_id': row.location_id,
        'plan_date': row.plan_date.isoformat() if row.plan_date else None,
        'dry_run': row.dry_run,
        'file_name': row.file_name,
        'plan_id': row.plan_id,
        'run_id': row.run_id,
    }


def report_detail(row: ExcelCompareReport) -> dict:
    data = report_summary(row)
    data['payload'] = row.payload
    return data
