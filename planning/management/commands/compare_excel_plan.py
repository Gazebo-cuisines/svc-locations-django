"""Draft-plan compare: Excel packing cases vs explode RM kg."""

from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from planning.errors import PlanningError, PlanningStateError
from planning.services.excel_compare import (
    ExcelCompareError,
    persist_report,
    run_excel_compare,
)


class Command(BaseCommand):
    help = 'Create a draft plan from Excel packing cases and compare RM kg.'

    def add_arguments(self, parser):
        parser.add_argument('--excel', required=True)
        parser.add_argument('--location-id', type=int, required=True)
        parser.add_argument('--plan-date', required=True, help='YYYY-MM-DD')
        parser.add_argument(
            '--qty-mode',
            choices=('packs', 'cases'),
            default='cases',
        )
        parser.add_argument('--out', default=None)
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--plan-id', type=int, default=None)

    def handle(self, *args, **options):
        excel_path = Path(options['excel']).expanduser().resolve()
        if not excel_path.exists():
            raise CommandError(f'file not found: {excel_path}')
        plan_date = parse_date(options['plan_date'])
        if plan_date is None:
            raise CommandError('--plan-date must be YYYY-MM-DD')
        try:
            result = run_excel_compare(
                source=excel_path,
                location_id=options['location_id'],
                plan_date=plan_date,
                qty_mode=options['qty_mode'],
                dry_run=options['dry_run'],
                plan_id=options['plan_id'],
                remarks=f'excel compare {excel_path.name}',
            )
        except ExcelCompareError as exc:
            raise CommandError(str(exc)) from exc
        except (PlanningError, PlanningStateError) as exc:
            raise CommandError(str(exc)) from exc

        pub = persist_report(result, excel_path.name)
        self.stdout.write(
            f'FG={len(result["finished_goods"])}  RM={len(result["rm_compare"])}  '
            f'plan={result["plan_id"]} run={result["run_id"]} '
            f'report_id={pub["report_id"]}'
        )
        for row in result['finished_goods']:
            self.stdout.write(
                f"  {row['excel_code']:18} cases={row['cases']} "
                f"qty={row['explode_qty']} → {row['product_id']} ({row['match']})"
            )
        out = Path(options['out']) if options['out'] else (
            excel_path.parent / f'compare-{excel_path.stem}.xlsx'
        )
        _write_report(out, result)
        self.stdout.write(self.style.SUCCESS(f'report {out}'))


_PCT_FORMAT = '+0.00"%";-0.00"%"'
_GREEN_FONT = Font(color='006100')
_RED_FONT = Font(color='9C0006')
_GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
_RED_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')


def _num(value):
    if value in (None, ''):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _colour_pct(sheet, cell_range: str):
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator='greaterThan',
            formula=['0'],
            font=_GREEN_FONT,
            fill=_GREEN_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator='lessThan',
            formula=['0'],
            font=_RED_FONT,
            fill=_RED_FILL,
        ),
    )


def _write_report(path: Path, result: dict):
    book = Workbook()
    fg_sheet = book.active
    fg_sheet.title = 'FG lines'
    fg_sheet.append([
        'excel_row', 'excel_code', 'excel_name', 'cases', 'packs',
        'explode_qty', 'product_id', 'recipe_code', 'product_name', 'match',
    ])
    for row in result['finished_goods']:
        fg_sheet.append([
            row['excel_row'], row['excel_code'], row['excel_name'],
            row['cases'], row['packs'], row['explode_qty'],
            row['product_id'], row['recipe_code'], row['product_name'],
            row['match'],
        ])
    cmp_sheet = book.create_sheet('RM compare')
    cmp_sheet.append([
        'excel_code', 'excel_name', 'excel_g', 'system_g',
        'diff', 'percentage', 'fix',
    ])
    for i, row in enumerate(result['rm_compare'], start=2):
        excel_g = _num(row.get('excel_g'))
        if excel_g is None:
            kg = _num(row['excel_kg'])
            excel_g = kg * 1000 if kg is not None else None
        system_g = _num(row.get('system_g'))
        if system_g is None:
            kg = _num(row['system_kg'])
            system_g = kg * 1000 if kg is not None else None
        cmp_sheet.cell(i, 1, row['excel_code'])
        cmp_sheet.cell(i, 2, row['excel_name'])
        cmp_sheet.cell(i, 3, excel_g)
        cmp_sheet.cell(i, 4, system_g)
        if excel_g:
            cmp_sheet.cell(i, 5, f'=D{i}-C{i}')
            pct = cmp_sheet.cell(i, 6, f'=(D{i}-C{i})/C{i}*100')
            pct.number_format = _PCT_FORMAT
        cmp_sheet.cell(i, 7, row['fix'])
    last = 1 + len(result['rm_compare'])
    if last >= 2:
        _colour_pct(cmp_sheet, f'F2:F{last}')
    extra = book.create_sheet('System only')
    extra.append(['product_id', 'recipe_code', 'name', 'system_kg', 'unit', 'fix'])
    for row in result['system_only']:
        extra.append([
            row['product_id'], row['recipe_code'], row['product_name'],
            row['system_kg'], row['unit'], row['fix'],
        ])
    meta = book.create_sheet('meta')
    meta.append(['plan_id', result['plan_id']])
    meta.append(['run_id', result['run_id']])
    meta.append(['dry_run', result['dry_run']])
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
