from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase
from openpyxl import Workbook, load_workbook

from locations.models import Location
from planning.management.commands.compare_excel_plan import _write_report
from planning.models import ExcelCompareReport
from planning.services.excel_compare import parse_packing


def _xlsx_bytes(rows):
    book = Workbook()
    sheet = book.active
    sheet.title = 'PACKING PLAN'
    for i, (code, name, cases, trays, per_case) in enumerate(rows, start=5):
        sheet.cell(i, 2, code)
        sheet.cell(i, 3, name)
        sheet.cell(i, 10, trays)
        sheet.cell(i, 13, cases)
        sheet.cell(i, 15, per_case)
    buf = BytesIO()
    book.save(buf)
    return buf.getvalue()


class ParsePackingTests(SimpleTestCase):
    def test_keeps_positive_cases_and_skips_zero(self):
        book = Workbook()
        sheet = book.active
        sheet.cell(5, 2, 'SKIP')
        sheet.cell(5, 3, 'Zero')
        sheet.cell(5, 13, 0)
        sheet.cell(6, 2, 'CVSAL-G12T')
        sheet.cell(6, 3, 'Veg samosa')
        sheet.cell(6, 10, 120)
        sheet.cell(6, 13, 10)
        sheet.cell(6, 15, 12)
        lines = parse_packing(sheet)
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]['code'], 'CVSAL-G12T')
        self.assertEqual(lines[0]['cases'], Decimal('10'))
        self.assertEqual(lines[0]['packs'], Decimal('120'))


class ExcelCompareApiTests(TestCase):
    def setUp(self):
        self.location = Location.objects.create(id=1, name='Grab and Go')

    def test_missing_file_is_400(self):
        res = self.client.post(
            '/planning/excel-compare/',
            {'location_id': self.location.id, 'plan_date': '2026-08-19'},
        )
        self.assertEqual(res.status_code, 400)

    def test_dry_run_parses_workbook(self):
        upload = SimpleUploadedFile(
            'day.xlsx',
            _xlsx_bytes([('CVSAL-G12T', 'Veg samosa', 10, 120, 12)]),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        res = self.client.post(
            '/planning/excel-compare/',
            {
                'location_id': str(self.location.id),
                'plan_date': '2026-08-19',
                'dry_run': 'true',
                'file': upload,
            },
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertEqual(body['status'], 'success')
        self.assertEqual(len(body['data']['finished_goods']), 1)
        self.assertTrue(body['data']['dry_run'])
        self.assertEqual(body['data']['qty_mode'], 'cases')
        self.assertIsNone(body['data']['plan_id'])
        report_id = body['data']['report_id']
        self.assertIsNotNone(report_id)
        self.assertEqual(ExcelCompareReport.objects.count(), 1)

        listed = self.client.get('/planning/excel-compare/')
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(len(listed.json()['data']['items']), 1)

        detail = self.client.get(f'/planning/excel-compare/{report_id}/')
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(len(detail.json()['data']['payload']['finished_goods']), 1)


class WriteReportPctTests(SimpleTestCase):
    def test_pct_formula_is_system_minus_excel_over_system(self):
        result = {
            'plan_id': 1,
            'run_id': 1,
            'dry_run': False,
            'finished_goods': [],
            'rm_compare': [{
                'excel_code': 'RMHAR002-01',
                'excel_name': 'PEAS (FROZEN)',
                'excel_kg': '387.5',
                'excel_g': '387500',
                'product_id': 1,
                'recipe_code': 'VEGFRO-01',
                'product_name': 'PEAS (FROZEN)',
                'match': 'map',
                'system_kg': '387.1',
                'system_g': '387100',
                'diff_kg': None,
                'diff_g': None,
                'diff_pct': None,
                'fix': None,
            }],
            'system_only': [],
            'explode_audit': [{
                'level': 5,
                'product_id': 7,
                'recipe_code': 'PASTRY-01',
                'product_name': 'SAMOSA PASTRY - LARGE CUT',
                'unit': 'grams',
                'kind': 'child',
                'summary': '17955.5616 from parent 579.96 × BOM 1.',
                'step': 2,
                'op': 'convert_uom',
                'formula': 'bom_unit → stock_unit',
                'from': '579.96 unit',
                'to': '17955.5616 grams',
                'skipped': '',
                'reason': '',
                'factor': '1 unit = 0.030960 kg (product_packaging)',
                'conversion': '579.96 × 0.030960 kg/unit ÷ 0.001000 kg/grams',
                'conversion_source': 'product_packaging',
                'result_net': '',
                'result_gross': '',
            }],
        }
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / 'compare.xlsx'
            _write_report(path, result)
            book = load_workbook(path)
            sheet = book['RM compare']
            explode = book['Explode']
            self.assertEqual(explode['O2'].value, '1 unit = 0.030960 kg (product_packaging)')
            self.assertEqual(explode['P2'].value, '579.96 × 0.030960 kg/unit ÷ 0.001000 kg/grams')
        self.assertEqual(sheet['C2'].value, 387500)
        self.assertEqual(sheet['D2'].value, 387100)
        self.assertEqual(sheet['E2'].value, '=D2-C2')
        self.assertEqual(sheet['F2'].value, '=(D2-C2)/D2*100')
        self.assertEqual(sheet['F2'].number_format, '+0.00"%";-0.00"%"')
        self.assertEqual(sheet['G2'].value, '=D2/C2*100')
        self.assertEqual(len(sheet.conditional_formatting._cf_rules), 1)
