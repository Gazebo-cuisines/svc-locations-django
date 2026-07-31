"""One-shot: explode legacy recipe tree for product 1227 → Excel."""
import os
from collections import deque
from decimal import Decimal
from pathlib import Path

import MySQLdb
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = 1227
OUT = Path(__file__).resolve().parent / "exports" / "product_1227_recipe_journey.xlsx"


def conn():
    load_dotenv()
    return MySQLdb.connect(
        host=os.getenv("DB_HOST"),
        port=int(os.getenv("DB_PORT") or 3306),
        user=os.getenv("DB_USER"),
        passwd=os.getenv("DB_PASSWORD"),
        db=os.getenv("LEGACY_DB_NAME", "production"),
        connect_timeout=30,
    )


def dec(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def fetch_products(cur, ids):
    if not ids:
        return {}
    ids = list(ids)
    ph = ",".join(["%s"] * len(ids))
    cur.execute(
        f"""
        SELECT id, productname, alternateproductname, productreceipecode,
               gffCode, productyield, productclass, `range`, subrange,
               categorypatth, srccontainer, destcontainer, unit, purchasingunit,
               unitcost, unitprice, casesizedescription, packweight,
               unitaryweight, grossunitaryweight, itemsperunit,
               containerVessel, tray, box, genhasreceipe, genPurchaseItem,
               genSalesItem, genrecordflag, remarks
        FROM tblproducts WHERE id IN ({ph})
        """,
        ids,
    )
    return {r["id"]: r for r in cur.fetchall()}


def fetch_containers(cur, ids):
    if not ids:
        return {}
    ids = list(ids)
    ph = ",".join(["%s"] * len(ids))
    cur.execute(f"SELECT * FROM tblcontainers WHERE id IN ({ph})", ids)
    return {r["id"]: r for r in cur.fetchall()}


def fetch_units(cur):
    cur.execute("SHOW TABLES LIKE 'units'")
    if not cur.fetchone():
        cur.execute("SHOW TABLES LIKE '%unit%'")
        tables = [list(t.values())[0] if isinstance(t, dict) else t[0] for t in cur.fetchall()]
        print("unit tables", tables)
    # try common names
    for table in ("Units", "units", "tblunits"):
        try:
            cur.execute(f"SELECT * FROM {table} LIMIT 1")
            sample = cur.fetchone()
            cur.execute(f"SELECT * FROM {table}")
            rows = cur.fetchall()
            key = "id" if "id" in (sample or {}) else list(sample.keys())[0]
            return {r[key]: r for r in rows}
        except Exception:
            continue
    return {}


def fetch_children(cur, parent_id):
    cur.execute(
        """
        SELECT id, active, version, stepinstructions, idx, parentprod, item,
               productyield, quantity, batchquantity, grossbatchquantity,
               unit, itemcost, linecost, implicit, inserttime, updatetime
        FROM tblproducttree
        WHERE parentprod=%s
        ORDER BY idx
        """,
        (parent_id,),
    )
    return list(cur.fetchall())


def fetch_npd_versions(cur, product_ids):
    cur.execute("SHOW COLUMNS FROM tblnpdproducttreeversion")
    cols = [r["Field"] for r in cur.fetchall()]
    # find product/parent column
    prod_col = None
    for c in ("product", "parentprod", "prod", "item"):
        if c in cols:
            prod_col = c
            break
    if not prod_col:
        return [], cols
    ids = list(product_ids)
    ph = ",".join(["%s"] * len(ids))
    cur.execute(
        f"SELECT * FROM tblnpdproducttreeversion WHERE `{prod_col}` IN ({ph})",
        ids,
    )
    return list(cur.fetchall()), cols


def fetch_npd_lines(cur, product_ids):
    ids = list(product_ids)
    ph = ",".join(["%s"] * len(ids))
    cur.execute(
        f"""
        SELECT * FROM tblnpdproducttree
        WHERE parentprod IN ({ph})
        ORDER BY parentprod, version, idx
        """,
        ids,
    )
    return list(cur.fetchall())


def explode_tree(cur, root_id):
    """BFS full BOM. Returns flat journey rows."""
    rows = []
    seen_edges = set()
    q = deque([(root_id, 0, "ROOT", None)])
    visited_parents = set()

    while q:
        parent_id, depth, path, parent_edge = q.popleft()
        if parent_id in visited_parents:
            continue
        visited_parents.add(parent_id)
        children = fetch_children(cur, parent_id)
        if not children and depth == 0:
            rows.append(
                {
                    "depth": 0,
                    "path": path,
                    "parent_id": None,
                    "component_id": root_id,
                    "line_no": None,
                    "quantity": None,
                    "unit_id": None,
                    "productyield": None,
                    "batchquantity": None,
                    "grossbatchquantity": None,
                    "itemcost": None,
                    "linecost": None,
                    "implicit": None,
                    "version": None,
                    "stepinstructions": None,
                    "tree_row_id": None,
                    "note": "no children in tblproducttree",
                }
            )
        for ch in children:
            edge = (parent_id, ch["item"], ch["idx"])
            if edge in seen_edges:
                continue
            seen_edges.add(edge)
            child_path = f"{path} > {ch['item']}"
            rows.append(
                {
                    "depth": depth + 1,
                    "path": child_path,
                    "parent_id": parent_id,
                    "component_id": ch["item"],
                    "line_no": ch["idx"],
                    "quantity": dec(ch["quantity"]),
                    "unit_id": ch["unit"],
                    "productyield": dec(ch["productyield"]),
                    "batchquantity": dec(ch["batchquantity"]),
                    "grossbatchquantity": dec(ch["grossbatchquantity"]),
                    "itemcost": dec(ch["itemcost"]),
                    "linecost": dec(ch["linecost"]),
                    "implicit": ch["implicit"],
                    "version": ch["version"],
                    "stepinstructions": ch["stepinstructions"],
                    "tree_row_id": ch["id"],
                    "note": "",
                }
            )
            q.append((ch["item"], depth + 1, child_path, edge))
    return rows, visited_parents


def sheet_write(ws, headers, rows):
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(h) for h in headers])


def main():
    c = conn()
    cur = c.cursor(MySQLdb.cursors.DictCursor)

    journey, parents = explode_tree(cur, ROOT)
    all_ids = {ROOT} | parents | {r["component_id"] for r in journey if r["component_id"]}
    products = fetch_products(cur, all_ids)

    container_ids = set()
    for p in products.values():
        for k in ("srccontainer", "destcontainer", "containerVessel"):
            if p.get(k) is not None:
                container_ids.add(p[k])
    containers = fetch_containers(cur, container_ids)
    units = fetch_units(cur)

    # enrich journey
    for r in journey:
        parent = products.get(r["parent_id"]) or {}
        child = products.get(r["component_id"]) or {}
        r["parent_name"] = parent.get("productname")
        r["component_name"] = child.get("productname")
        r["component_code"] = child.get("productreceipecode")
        r["component_yield"] = dec(child.get("productyield"))
        r["component_has_recipe"] = child.get("genhasreceipe")
        r["component_src_container_id"] = child.get("srccontainer")
        r["component_dest_container_id"] = child.get("destcontainer")
        src = containers.get(child.get("srccontainer") or -1) or {}
        dst = containers.get(child.get("destcontainer") or -1) or {}
        # container display: try common name fields
        def cname(row):
            for k in ("container", "name", "description", "containername"):
                if k in row and row[k]:
                    return row[k]
            return None

        r["container_in_id"] = child.get("srccontainer")
        r["container_in_name"] = cname(src)
        r["container_out_id"] = child.get("destcontainer")
        r["container_out_name"] = cname(dst)
        u = units.get(r["unit_id"]) or {}
        r["unit_name"] = u.get("unit") or u.get("name") or u.get("Unit")

    # product master sheet
    product_rows = []
    for pid, p in sorted(products.items()):
        src = containers.get(p.get("srccontainer") or -1) or {}
        dst = containers.get(p.get("destcontainer") or -1) or {}

        def cname(row):
            for k in ("container", "name", "description", "containername"):
                if k in row and row[k]:
                    return row[k]
            return None

        product_rows.append(
            {
                "id": p["id"],
                "productname": p.get("productname"),
                "alternateproductname": p.get("alternateproductname"),
                "productreceipecode": p.get("productreceipecode"),
                "gffCode": p.get("gffCode"),
                "productyield": dec(p.get("productyield")),
                "productclass": p.get("productclass"),
                "range": p.get("range"),
                "subrange": p.get("subrange"),
                "categorypatth": p.get("categorypatth"),
                "container_in_id": p.get("srccontainer"),
                "container_in_name": cname(src),
                "container_out_id": p.get("destcontainer"),
                "container_out_name": cname(dst),
                "containerVessel": p.get("containerVessel"),
                "unit": p.get("unit"),
                "purchasingunit": p.get("purchasingunit"),
                "unitcost": dec(p.get("unitcost")),
                "unitprice": dec(p.get("unitprice")),
                "casesizedescription": p.get("casesizedescription"),
                "packweight": dec(p.get("packweight")),
                "unitaryweight": dec(p.get("unitaryweight")),
                "grossunitaryweight": dec(p.get("grossunitaryweight")),
                "itemsperunit": p.get("itemsperunit"),
                "tray": p.get("tray"),
                "box": p.get("box"),
                "genhasreceipe": p.get("genhasreceipe"),
                "genPurchaseItem": p.get("genPurchaseItem"),
                "genSalesItem": p.get("genSalesItem"),
                "genrecordflag": p.get("genrecordflag"),
                "remarks": p.get("remarks"),
            }
        )

    # containers sheet (full dump of used)
    container_rows = []
    if containers:
        # pick useful cols
        sample = next(iter(containers.values()))
        prefer = [
            c
            for c in (
                "id",
                "container",
                "name",
                "description",
                "active",
                "topcontainer",
                "parentcontainer",
                "remarks",
            )
            if c in sample
        ]
        extra = [c for c in sample.keys() if c not in prefer][:25]
        cheaders = prefer + extra
        for cid, row in sorted(containers.items()):
            container_rows.append({h: row.get(h) for h in cheaders})
    else:
        cheaders = ["id"]

    npd_vers, _ = fetch_npd_versions(cur, all_ids)
    npd_lines = fetch_npd_lines(cur, all_ids)

    # root header sheet
    root = products.get(ROOT) or {}
    src = containers.get(root.get("srccontainer") or -1) or {}
    dst = containers.get(root.get("destcontainer") or -1) or {}

    def cname(row):
        for k in ("container", "name", "description", "containername"):
            if k in row and row[k]:
                return row[k]
        return None

    wb = Workbook()

    # 1) Overview
    ws = wb.active
    ws.title = "01_overview"
    overview = [
        ("product_id", ROOT),
        ("productname", root.get("productname")),
        ("productreceipecode", root.get("productreceipecode")),
        ("productyield", dec(root.get("productyield"))),
        ("container_in_id (src)", root.get("srccontainer")),
        ("container_in_name", cname(src)),
        ("container_out_id (dest)", root.get("destcontainer")),
        ("container_out_name", cname(dst)),
        ("has_recipe", root.get("genhasreceipe")),
        ("casesizedescription", root.get("casesizedescription")),
        ("packweight", dec(root.get("packweight"))),
        ("unitaryweight", dec(root.get("unitaryweight"))),
        ("nodes_in_tree", len(all_ids)),
        ("direct_components", sum(1 for r in journey if r["depth"] == 1)),
        ("total_bom_lines", len(journey)),
        ("source", "production.tblproducttree + tblproducts + tblcontainers"),
    ]
    ws.append(["field", "value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for a, b in overview:
        ws.append([a, b])

    # 2) Journey BOM
    jheaders = [
        "depth",
        "path",
        "parent_id",
        "parent_name",
        "line_no",
        "component_id",
        "component_name",
        "component_code",
        "quantity",
        "unit_id",
        "unit_name",
        "productyield",
        "component_yield",
        "batchquantity",
        "grossbatchquantity",
        "itemcost",
        "linecost",
        "implicit",
        "version",
        "component_has_recipe",
        "container_in_id",
        "container_in_name",
        "container_out_id",
        "container_out_name",
        "stepinstructions",
        "tree_row_id",
        "note",
    ]
    ws2 = wb.create_sheet("02_bom_journey")
    sheet_write(ws2, jheaders, journey)

    # 3) Products
    pheaders = list(product_rows[0].keys()) if product_rows else ["id"]
    ws3 = wb.create_sheet("03_products")
    sheet_write(ws3, pheaders, product_rows)

    # 4) Containers
    ws4 = wb.create_sheet("04_containers")
    sheet_write(ws4, cheaders, container_rows)

    # 5) NPD versions
    ws5 = wb.create_sheet("05_npd_versions")
    if npd_vers:
        nh = list(npd_vers[0].keys())
        sheet_write(ws5, nh, [{k: dec(v) if isinstance(v, Decimal) else v for k, v in r.items()} for r in npd_vers])
    else:
        ws5.append(["note"])
        ws5.append(["No NPD version rows for products in this tree"])

    # 6) NPD lines
    ws6 = wb.create_sheet("06_npd_tree_lines")
    if npd_lines:
        nh = list(npd_lines[0].keys())
        sheet_write(ws6, nh, [{k: dec(v) if isinstance(v, Decimal) else v for k, v in r.items()} for r in npd_lines])
    else:
        ws6.append(["note"])
        ws6.append(["No NPD tree lines for products in this tree"])

    # 7) Container flow (parent dest → child src style summary)
    ws7 = wb.create_sheet("07_container_flow")
    flow_headers = [
        "depth",
        "parent_id",
        "parent_name",
        "parent_container_out_id",
        "parent_container_out_name",
        "component_id",
        "component_name",
        "component_container_in_id",
        "component_container_in_name",
        "component_container_out_id",
        "component_container_out_name",
        "quantity",
        "unit_id",
    ]
    flow_rows = []
    for r in journey:
        parent = products.get(r["parent_id"]) or {}
        pdst = containers.get(parent.get("destcontainer") or -1) or {}
        flow_rows.append(
            {
                "depth": r["depth"],
                "parent_id": r["parent_id"],
                "parent_name": r.get("parent_name"),
                "parent_container_out_id": parent.get("destcontainer"),
                "parent_container_out_name": cname(pdst),
                "component_id": r["component_id"],
                "component_name": r.get("component_name"),
                "component_container_in_id": r.get("container_in_id"),
                "component_container_in_name": r.get("container_in_name"),
                "component_container_out_id": r.get("container_out_id"),
                "component_container_out_name": r.get("container_out_name"),
                "quantity": r.get("quantity"),
                "unit_id": r.get("unit_id"),
            }
        )
    # also root row
    flow_rows.insert(
        0,
        {
            "depth": 0,
            "parent_id": None,
            "parent_name": None,
            "parent_container_out_id": None,
            "parent_container_out_name": None,
            "component_id": ROOT,
            "component_name": root.get("productname"),
            "component_container_in_id": root.get("srccontainer"),
            "component_container_in_name": cname(src),
            "component_container_out_id": root.get("destcontainer"),
            "component_container_out_name": cname(dst),
            "quantity": None,
            "unit_id": root.get("unit"),
        },
    )
    sheet_write(ws7, flow_headers, flow_rows)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"products={len(products)} bom_lines={len(journey)} containers={len(containers)}")
    print("Direct children:")
    for r in journey:
        if r["depth"] == 1:
            print(f"  [{r['line_no']}] {r['component_id']} {r['component_name']} qty={r['quantity']}")
    c.close()


if __name__ == "__main__":
    main()
